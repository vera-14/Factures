# -*- coding: utf-8 -*-
from FormingDB import *
from MeterRhyme.PoemAnalyser.DBReaderFile import DBReader
from MeterRhyme.PoemAnalyser.MeterAnalyzer2 import MeterAnalyzer1
from MeterRhyme.PoemAnalyser.RhymeAnalyser2 import *
import openpyxl as opxl


def ExcelPushkin():
    dbreader = DBReader({
        'dbtype': 'sqlite',
        'path': '/Users/veram/PycharmProjects/MeterRhyme/database/accents_new.db'})
    dbreader.connect()
    acc = Accentuation(dbreader)
    meter = MeterAnalyzer1()
    rhyme = RhymeAnalyser()
    listOfStih = readText("/Users/veram/PycharmProjects/Factures/MeterRhyme/pushkinCorpus.txt")
    table = opxl.load_workbook('results.xlsx')
    sheet = table.active
    n = 0
    for Stih in listOfStih:
        poemstihdata = acc.analyze2(Stih)
        poemmeter = meter.Hello(poemstihdata)
        poemstrophica = rhyme.rhyme(Stih.getCleanLines(), poemstihdata)
        n = n+1
        cell = sheet.cell(row=n, column=1)
        cell.value = Stih.name
        cell = sheet.cell(row=n, column=2)
        cell.value = poemstrophica[1]
        cell = sheet.cell(row=n, column=3)
        cell.value = poemmeter
        cell = sheet.cell(row=n, column=4)
        cell.value = meter.stopCount()
        if poemstrophica[1] != 'свободная':
            print('\nПараметры входного стихотворения:', Stih.name, poemstrophica[1], poemmeter, meter.stopCount())

            sql1 = "SELECT * FROM samples WHERE strophica=? and meter=? and stopnost=?"
            cur.execute(sql1, (poemstrophica[1], poemmeter, meter.stopCount()))
            result = cur.fetchall()
            if result == []:
                print('Фактура не найдена')

            else:
                print('Номер соответствующей фактуры:', result[0][0])
                print('Пример строфы с такой же фактурой:\n', result[0][7], '\n', result[0][4], '\n', '№', result[0][5],
                      ',', result[0][6], 'год')
                cell = sheet.cell(row=n, column=5)
                cell.value = result[0][0]
                cell = sheet.cell(row=n, column=6)
                cell.value = result[0][7]
                cell = sheet.cell(row=n, column=7)
                cell.value = result[0][4]
                cell = sheet.cell(row=n, column=8)
                cell.value = result[0][5]
                cell = sheet.cell(row=n, column=9)
                cell.value = result[0][6]
        table.save('results.xlsx')


def ExcelTolstoy():
    dbreader = DBReader({
        'dbtype': 'sqlite',
        'path': '/Users/veram/PycharmProjects/MeterRhyme/database/accents_new.db'})
    dbreader.connect()
    acc = Accentuation(dbreader)
    meter = MeterAnalyzer1()
    rhyme = RhymeAnalyser()
    listOfStih = readText("/Users/veram/PycharmProjects/Factures/MeterRhyme/Tolstoy1.txt")
    table = opxl.load_workbook('resultsTolstoy.xlsx')
    sheet = table.active
    n = 0
    for Stih in listOfStih:
        poemstihdata = acc.analyze2(Stih)
        poemmeter = meter.Hello(poemstihdata)
        poemstrophica = rhyme.rhyme(Stih.getCleanLines(), poemstihdata)
        n = n + 1
        cell = sheet.cell(row=n, column=1)
        cell.value = Stih.name
        cell = sheet.cell(row=n, column=2)
        cell.value = poemstrophica[1]
        cell = sheet.cell(row=n, column=3)
        cell.value = poemmeter
        cell = sheet.cell(row=n, column=4)
        cell.value = meter.stopCount()
        if poemstrophica[1] != 'свободная':
            #print('\nПараметры входного стихотворения:', Stih.name, poemstrophica[1], poemmeter, meter.stopCount())

            sql1 = "SELECT * FROM samples WHERE strophica=? and meter=? and stopnost=?"
            cur.execute(sql1, (poemstrophica[1], poemmeter, meter.stopCount()))
            result = cur.fetchall()
            if result == []:
                print(Stih.name)
                #print('Фактура не найдена')

            else:
                #print('Номер соответствующей фактуры:', result[0][0])
                #print('Пример строфы с такой же фактурой:\n', result[0][7], '\n', result[0][4], '\n', '№', result[0][5],
                      #',', result[0][6], 'год')
                cell = sheet.cell(row=n, column=5)
                cell.value = result[0][0]
                cell = sheet.cell(row=n, column=6)
                cell.value = result[0][7]
                cell = sheet.cell(row=n, column=7)
                cell.value = result[0][4]
                cell = sheet.cell(row=n, column=8)
                cell.value = result[0][5]
                cell = sheet.cell(row=n, column=9)
                cell.value = result[0][6]
                print(Stih.name)
        table.save('resultsTolstoy.xlsx')
